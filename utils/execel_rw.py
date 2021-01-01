import openpyxl
from openpyxl.styles.fonts import Font

class Execel_rw():
    def __init__(self, file, sheet):
        self.wb = openpyxl.load_workbook(file)
        self.ws = self.wb[sheet]
        self.file = file
        self.red = Font('微软雅黑', bold=True, color='FF0000')
        self.green = Font('微软雅黑', bold=True, color='00FF00')

    def cell(self, row, col, value=None):
        return self.ws.cell(row, col, value)

    def rows(self):
        return self.ws.rows

    def max_row(self):
        return self.ws.max_row

    def insert_rows(self,idx,amount):
        self.ws.insert_rows(idx, amount)

    def insert_cols(self, idx, amount):
        self.ws.insert_cols(idx, amount)

    def delete_rows(self, idx, amount):
        self.ws.delete_rows(idx, amount)

    def delte_cols(self, idx, amount):
        self.ws.delete_cols(idx, amount)

    def save(self):
        self.wb.save(self.file)



# print(type(Execel_rw('api.xlsx','Sheet').cell(2,1).value))
# data = Execel_rw('api.xlsx','Sheet1')
# print(data.total_row())
# table = data.get_sheet_by_name('Sheet')

# nrows = table.rows # 获得行数类型为迭代器
# ncols = table.columns
# print(type(nrows)) # generator

# for row in nrows:
#     print(row) # (<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>)
#     line = [col.value for col in row] # 取值
#     for col in row:
#         print(col.value)

# print(table.cell(1,2).value)

# print(table)
# ws = wb.active
# ws['A1']='name'
# ws['B1']='url'
# wb.save('api.xlsx')

# print(data.get_named_ranges())
# print(data.get_sheet_names())
# sheet_names = data.get_sheet_names()
# table = data.get_sheet_by_name(sheet_names[0])
# table = data.active
# print(table.title)
# nrows = table.max_row
# ncolumns = table.max_column
# nrows = table.rows
# vars = ['name','url','method','header','body','expected','response','result']
# table.cell(1,2,'url')
# table.append(vars)
# table.delete_rows(1,2)
# data.save('api.xlsx')